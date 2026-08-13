#!/usr/bin/env python3
"""Create separate state-wise Excel workbooks from CSV files in Google Drive.

Google Drive authentication, exact-name lookup, downloading, and CSV reading follow
the same structure as the original script. Each source CSV contains one data table;
no input worksheet name is configured or used.
"""

import io
import json
import logging
import os
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

import pandas as pd
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from googleapiclient.http import MediaIoBaseDownload
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DRIVE_SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]

# Same configuration style as the old code: tool name, exact Drive filename, reader.
# No input sheet name is required because all source files are CSV files.
TOOL_FILES = [
    {"tool_name": "DOD", "filename": "DOD_2026_WIDE.csv", "reader": "csv"},
    {"tool_name": "Maternal_Log", "filename": "Maternal Log_WIDE.csv", "reader": "csv"},
    {"tool_name": "FIS", "filename": "FIS CASE RECORDS 2026_WIDE.csv", "reader": "csv"},
    {"tool_name": "PMSMA", "filename": "PMSMA Client Interview_WIDE.csv", "reader": "csv"},
    {"tool_name": "SNCU", "filename": "SNCU Index Cases Observation 2026_WIDE.csv", "reader": "csv"},
    {"tool_name": "Referral_Services", "filename": "GABV IDD Referral_WIDE.csv", "reader": "csv"},
    {"tool_name": "Digital_System", "filename": "GAVB IDD Digital System_WIDE.csv", "reader": "csv"},
    {"tool_name": "Exit_Interview", "filename": "GAVB IDD Exit Interview_WIDE.csv", "reader": "csv"},
    {"tool_name": "HR", "filename": "GAVB IDD HR_WIDE.csv", "reader": "csv"},
    {"tool_name": "Labour_Room_Readiness", "filename": "GAVB IDD Labour Room Readiness_WIDE.csv", "reader": "csv"},
    {"tool_name": "Supply_Chain", "filename": "GAVB IDD Supply Chain_WIDE.csv", "reader": "csv"},
]

TOOL_LABELS = {
    "DOD": "DOD",
    "Maternal_Log": "Maternal Log",
    "FIS": "FIS",
    "PMSMA": "PMSMA",
    "SNCU": "SNCU",
    "Referral_Services": "Referral Services",
    "Digital_System": "Digital System",
    "Exit_Interview": "Exit Interview",
    "HR": "HR",
    "Labour_Room_Readiness": "Labour Room Readiness",
    "Supply_Chain": "Supply Chain",
}
TOOL_ORDER = [item["tool_name"] for item in TOOL_FILES]

COLUMN_ALIASES = {
    "state": ["STATE", "State", "state", "Cal_STATE", "state_name"],
    "district": ["Cal_DIST", "District", "DISTRICT", "district", "district_name"],
    "investigator": [
        "QDC", "Investigator", "Nurse", "Nurse_Name", "Nursing_Consultant",
        "Name of Nursing Consultants", "Name of Nurses", "collector_name",
    ],
    "facility_type": ["F_Type", "Facility_Type", "Facility Type", "facilitytype"],
    "facility_level": ["Facility_Level", "Facility Level", "Level", "DH_Below_DH"],
    "submission_date": [
        "SubmissionDate", "Submission Date", "submission_date",
        "SubmissionDateTime", "Submission_Time", "starttime", "endtime",
    ],
}

DH_VALUES = {"dh", "district hospital", "district_hospital", "district hospital dh"}
DAY_START_HOUR = 9
DAY_END_HOUR = 18
DEFAULT_STATE_SPOC = {"Assam": "Nikhil Kumar"}

HEADER_FILL = PatternFill("solid", fgColor="B7DEE8")
TOTAL_FILL = PatternFill("solid", fgColor="B7DEE8")
THIN = Side(style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# =============================================================================
# ORIGINAL GOOGLE DRIVE READING STRUCTURE
# =============================================================================
def escape_drive_query_value(value: str) -> str:
    return value.replace("\\", "\\\\").replace("'", "\\'")


def required_env(name: str) -> str:
    value = os.getenv(name, "").strip()
    if not value:
        raise ValueError(f"Missing required environment variable: {name}")
    return value


def build_drive_credentials() -> Credentials:
    service_account_json = os.getenv("GOOGLE_SERVICE_ACCOUNT_JSON", "").strip()
    service_account_file = os.getenv("GOOGLE_SERVICE_ACCOUNT_FILE", "credential.json").strip()

    if service_account_json:
        try:
            credentials_info = json.loads(service_account_json)
        except json.JSONDecodeError as exc:
            raise ValueError("GOOGLE_SERVICE_ACCOUNT_JSON is not valid JSON") from exc
        return Credentials.from_service_account_info(credentials_info, scopes=DRIVE_SCOPES)

    if os.path.exists(service_account_file):
        return Credentials.from_service_account_file(service_account_file, scopes=DRIVE_SCOPES)

    raise ValueError(
        "Service account credentials not found. Set GOOGLE_SERVICE_ACCOUNT_JSON "
        "or provide GOOGLE_SERVICE_ACCOUNT_FILE (default: credential.json)."
    )


def find_file_id_by_name(drive_service, folder_id: str, filename: str) -> str:
    safe_name = escape_drive_query_value(filename)
    query = (
        f"'{folder_id}' in parents and trashed = false and "
        f"name = '{safe_name}'"
    )
    response = (
        drive_service.files()
        .list(q=query, spaces="drive", fields="files(id,name)", pageSize=10)
        .execute()
    )
    files = response.get("files", [])
    if not files:
        raise FileNotFoundError(f"File not found in Drive folder: {filename}")
    return files[0]["id"]


def download_drive_file_bytes(drive_service, file_id: str) -> bytes:
    request = drive_service.files().get_media(fileId=file_id)
    file_buffer = io.BytesIO()
    downloader = MediaIoBaseDownload(file_buffer, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    return file_buffer.getvalue()


def read_tool_file(data: bytes, reader: str) -> pd.DataFrame:
    data_stream = io.BytesIO(data)
    if reader == "csv":
        # Same byte-stream reading pattern as the old code. utf-8-sig also reads
        # normal UTF-8 and safely removes a BOM when one is present.
        return pd.read_csv(
            data_stream,
            low_memory=False,
            encoding="utf-8-sig",
        )
    raise ValueError(f"Unsupported reader: {reader}")


# =============================================================================
# STANDARDIZATION AND DATE PARSING
# =============================================================================
def normalize(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value).strip().lower())


def clean_text(series: pd.Series) -> pd.Series:
    return series.astype("string").str.strip().replace(
        {"": pd.NA, "nan": pd.NA, "None": pd.NA}
    )


def find_column(columns: Sequence[object], aliases: Sequence[str]) -> Optional[str]:
    lookup = {normalize(column): str(column) for column in columns}
    for alias in aliases:
        key = normalize(alias)
        if key in lookup:
            return lookup[key]
    return None


def parse_submission_datetime(series: pd.Series) -> pd.Series:
    """Parse SubmissionDate values such as 27/01/2026, 19:01:46."""
    text = clean_text(series)
    parsed = pd.to_datetime(
        text,
        format="%d/%m/%Y, %H:%M:%S",
        errors="coerce",
    )
    unresolved = parsed.isna() & text.notna()
    if unresolved.any():
        try:
            fallback = pd.to_datetime(
                text.loc[unresolved], format="mixed", dayfirst=True, errors="coerce"
            )
        except TypeError:
            fallback = pd.to_datetime(
                text.loc[unresolved], dayfirst=True, errors="coerce"
            )
        parsed.loc[unresolved] = fallback
    return parsed


def standardize_frame(
    raw: pd.DataFrame, tool_name: str, filename: str
) -> Tuple[pd.DataFrame, Dict[str, object]]:
    detected = {
        key: find_column(raw.columns, aliases)
        for key, aliases in COLUMN_ALIASES.items()
    }

    if not detected["state"] or not detected["investigator"]:
        raise ValueError(
            f"{filename}: missing State or Investigator/QDC column. "
            f"Available headers: {list(raw.columns)}"
        )

    frame = raw.copy()
    frame["__State"] = clean_text(frame[detected["state"]])
    frame["__Investigator"] = clean_text(frame[detected["investigator"]])

    level_source = detected["facility_level"] or detected["facility_type"]
    if level_source:
        dh_values = {normalize(value) for value in DH_VALUES}
        frame["__FacilityLevel"] = clean_text(frame[level_source]).map(
            lambda value: "DH" if normalize(value) in dh_values else "Below DH"
        )
    else:
        frame["__FacilityLevel"] = "Unclassified"

    if detected["submission_date"]:
        frame["__SubmissionDateTime"] = parse_submission_datetime(
            frame[detected["submission_date"]]
        )
    else:
        frame["__SubmissionDateTime"] = pd.NaT

    frame["__Tool"] = tool_name

    log = {
        "Tool": TOOL_LABELS[tool_name],
        "File": filename,
        "Rows_Read": len(raw),
        "Rows_With_State": int(frame["__State"].notna().sum()),
        "Rows_Without_State": int(frame["__State"].isna().sum()),
        "Rows_Without_Investigator": int(frame["__Investigator"].isna().sum()),
        "State_Column": detected["state"],
        "Investigator_Column": detected["investigator"],
        "Facility_Level_Column": level_source or "",
        "Submission_Date_Column": detected["submission_date"] or "",
        "Invalid_Submission_Dates": (
            int(frame["__SubmissionDateTime"].isna().sum())
            if detected["submission_date"] else "Not supplied"
        ),
        "Status": "OK",
    }
    return frame, log


def load_and_standardize_data(
    drive_service, source_folder_id: str
) -> Tuple[Dict[str, pd.DataFrame], pd.DataFrame]:
    frames: Dict[str, pd.DataFrame] = {}
    logs: List[Dict[str, object]] = []

    for item in TOOL_FILES:
        tool_name = item["tool_name"]
        filename = item["filename"]
        logging.info("Reading %s", filename)
        try:
            file_id = find_file_id_by_name(drive_service, source_folder_id, filename)
            raw_bytes = download_drive_file_bytes(drive_service, file_id)
            raw_frame = read_tool_file(raw_bytes, item["reader"])
            frame, log = standardize_frame(raw_frame, tool_name, filename)
            frames[tool_name] = frame
            logs.append(log)
        except Exception as exc:
            logging.exception("Could not process %s", filename)
            logs.append({
                "Tool": TOOL_LABELS[tool_name],
                "File": filename,
                "Rows_Read": 0,
                "Status": f"ERROR: {exc}",
            })

    if not frames:
        raise ValueError("No Google Drive CSV file could be processed.")
    return frames, pd.DataFrame(logs)


# =============================================================================
# REPORT TABLES
# =============================================================================
def filter_state(frame: pd.DataFrame, state: str) -> pd.DataFrame:
    return frame.loc[
        frame["__State"].fillna("").str.casefold() == state.casefold()
    ].copy()


def get_states(frames: Dict[str, pd.DataFrame]) -> List[str]:
    states = {
        str(value).strip()
        for frame in frames.values()
        for value in frame["__State"].dropna().unique()
        if str(value).strip()
    }
    return sorted(states, key=str.casefold)


def get_investigators(state_frames: Dict[str, pd.DataFrame]) -> List[str]:
    investigators = {
        str(value).strip()
        for frame in state_frames.values()
        for value in frame["__Investigator"].dropna().unique()
        if str(value).strip()
    }
    return sorted(investigators, key=str.casefold)


def count_by_investigator(
    frame: Optional[pd.DataFrame], investigators: List[str]
) -> List[int]:
    if frame is None or frame.empty:
        return [0] * len(investigators)
    grouped = frame.dropna(subset=["__Investigator"]).groupby("__Investigator").size()
    lookup = {
        str(name).strip().casefold(): int(count)
        for name, count in grouped.items()
    }
    return [lookup.get(name.casefold(), 0) for name in investigators]


def build_nurse_wise(
    state_frames: Dict[str, pd.DataFrame], investigators: List[str]
) -> pd.DataFrame:
    report = pd.DataFrame({"Name of Nursing Consultants": investigators})
    for tool_name in TOOL_ORDER:
        report[f"# {TOOL_LABELS[tool_name]}"] = count_by_investigator(
            state_frames.get(tool_name), investigators
        )
    total = {report.columns[0]: "Grand Total"}
    total.update({column: int(report[column].sum()) for column in report.columns[1:]})
    return pd.concat([report, pd.DataFrame([total])], ignore_index=True)


def build_dh_below_dh(
    state_frames: Dict[str, pd.DataFrame], investigators: List[str]
) -> pd.DataFrame:
    data: Dict[Tuple[str, str], List[object]] = {
        ("", "Name of Nurses"): investigators
    }
    for tool_name in TOOL_ORDER:
        frame = state_frames.get(tool_name)
        for level in ("Below DH", "DH"):
            subset = None if frame is None else frame.loc[
                frame["__FacilityLevel"] == level
            ]
            data[(TOOL_LABELS[tool_name], level)] = count_by_investigator(
                subset, investigators
            )
    report = pd.DataFrame(data)
    total = {report.columns[0]: "Grand Total"}
    total.update({column: int(report[column].sum()) for column in report.columns[1:]})
    return pd.concat([report, pd.DataFrame([total])], ignore_index=True)


def fis_shift_counts(
    frame: Optional[pd.DataFrame], investigators: List[str], day: bool
) -> List[int]:
    if frame is None or frame.empty:
        return [0] * len(investigators)
    hours = frame["__SubmissionDateTime"].dt.hour
    day_mask = (hours >= DAY_START_HOUR) & (hours < DAY_END_HOUR)
    subset = frame.loc[day_mask if day else (~day_mask & hours.notna())]
    return count_by_investigator(subset, investigators)


def build_summary(
    state: str,
    state_frames: Dict[str, pd.DataFrame],
    investigators: List[str],
    spoc_map: Dict[str, str],
) -> pd.DataFrame:
    report = pd.DataFrame({
        "Name of Nursing Consultants": investigators,
        "State": [state] * len(investigators),
        "State SPOC": [spoc_map.get(state, "")] * len(investigators),
    })
    for tool_name in TOOL_ORDER:
        report[TOOL_LABELS[tool_name]] = count_by_investigator(
            state_frames.get(tool_name), investigators
        )

    fis_position = report.columns.get_loc("FIS") + 1
    report.insert(
        fis_position,
        "FIS-Day (9AM-6PM)",
        fis_shift_counts(state_frames.get("FIS"), investigators, day=True),
    )
    report.insert(
        fis_position + 1,
        "FIS-Night (6PM-9AM)",
        fis_shift_counts(state_frames.get("FIS"), investigators, day=False),
    )
    return report


# =============================================================================
# EXCEL EXPORT
# =============================================================================
def safe_filename(value: str) -> str:
    return re.sub(r'[<>:"/\\|?*]+', "_", value).strip() or "Unknown_State"


def source_sheet_name(filename: str) -> str:
    """Derive the output raw-data sheet name from filename, not input metadata."""
    stem = Path(filename).stem
    cleaned = re.sub(r'[\\/*?:\[\]]', "_", stem).strip()
    return cleaned[:31] or "Raw_Data"


def raw_for_export(frame: pd.DataFrame) -> pd.DataFrame:
    helper_columns = [
        column for column in frame.columns if str(column).startswith("__")
    ]
    return frame.drop(columns=helper_columns, errors="ignore")


def autosize_worksheet(worksheet, maximum: int = 40) -> None:
    for column_index in range(1, worksheet.max_column + 1):
        letter = get_column_letter(column_index)
        max_length = 0
        for cell in worksheet[letter]:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[letter].width = min(max(max_length + 2, 10), maximum)


def style_workbook(output_path: Path) -> None:
    workbook = load_workbook(output_path)
    for worksheet in workbook.worksheets:
        if worksheet.title == "Summary":
            header_rows = (3,)
            worksheet.freeze_panes = "A4"
        elif worksheet.title == "DH & Below DH":
            header_rows = (1, 2)
            worksheet.freeze_panes = "B3"
        else:
            header_rows = (1,)
            worksheet.freeze_panes = "A2"

        for header_row in header_rows:
            for cell in worksheet[header_row]:
                cell.fill = HEADER_FILL
                cell.font = Font(bold=True)
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
                cell.border = BORDER

        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = BORDER

        if worksheet.title in {"Nurse Wise", "DH & Below DH"}:
            for cell in worksheet[worksheet.max_row]:
                cell.fill = TOTAL_FILL
                cell.font = Font(bold=True)

        autosize_worksheet(worksheet)
    workbook.save(output_path)


def create_state_workbook(
    state: str,
    frames: Dict[str, pd.DataFrame],
    processing_log: pd.DataFrame,
    output_folder: Path,
    spoc_map: Dict[str, str],
) -> Optional[Path]:
    state_frames = {
        tool_name: filter_state(frame, state)
        for tool_name, frame in frames.items()
    }
    investigators = get_investigators(state_frames)
    if not investigators:
        logging.warning("Skipping %s because no investigator was found", state)
        return None

    output_folder.mkdir(parents=True, exist_ok=True)
    report_date = datetime.now().strftime("%d-%m-%Y")
    output_path = output_folder / (
        f"{safe_filename(state)}_Tool_Submission_Report_{report_date}.xlsx"
    )

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        build_nurse_wise(state_frames, investigators).to_excel(
            writer, sheet_name="Nurse Wise", index=False
        )
        build_dh_below_dh(state_frames, investigators).to_excel(
            writer, sheet_name="DH & Below DH", index=True, merge_cells=True
        )
        build_summary(state, state_frames, investigators, spoc_map).to_excel(
            writer, sheet_name="Summary", index=False, startrow=2
        )

        # Raw-data output sheet names are automatically derived from CSV filenames.
        for item in TOOL_FILES:
            tool_name = item["tool_name"]
            if tool_name in state_frames:
                raw_for_export(state_frames[tool_name]).to_excel(
                    writer,
                    sheet_name=source_sheet_name(item["filename"]),
                    index=False,
                )

        processing_log.assign(State_Workbook=state).to_excel(
            writer, sheet_name="Processing Log", index=False
        )

    workbook = load_workbook(output_path)
    workbook["DH & Below DH"].delete_cols(1)
    workbook["Summary"]["A1"] = (
        f"GAVB Facility Tool Data collection status as on {report_date}"
    )
    workbook.save(output_path)
    style_workbook(output_path)
    return output_path


def parse_state_spoc() -> Dict[str, str]:
    mapping = DEFAULT_STATE_SPOC.copy()
    raw = os.getenv("STATE_SPOC_JSON", "").strip()
    if raw:
        supplied = json.loads(raw)
        if not isinstance(supplied, dict):
            raise ValueError("STATE_SPOC_JSON must be a JSON object")
        mapping.update({str(key).strip(): str(value).strip() for key, value in supplied.items()})
    return mapping


def main() -> int:
    logging.basicConfig(
        level=os.getenv("LOG_LEVEL", "INFO").upper(),
        format="%(asctime)s [%(levelname)s] %(message)s",
    )
    try:
        source_folder_id = required_env("DRIVE_FOLDER_ID")
        run_date_name = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        output_folder = Path(
            os.getenv("LOCAL_OUTPUT_FOLDER", f"/tmp/{run_date_name}/state_reports")
        )

        credentials = build_drive_credentials()
        drive_service = build(
            "drive", "v3", credentials=credentials, cache_discovery=False
        )

        frames, processing_log = load_and_standardize_data(
            drive_service, source_folder_id
        )
        states = get_states(frames)

        selected_states = os.getenv("STATES", "").strip()
        if selected_states:
            wanted = {
                value.strip().casefold()
                for value in selected_states.split(",")
                if value.strip()
            }
            states = [state for state in states if state.casefold() in wanted]

        if not states:
            raise ValueError("No matching State values were found.")

        spoc_map = parse_state_spoc()
        created: List[Path] = []
        for state in states:
            output_path = create_state_workbook(
                state, frames, processing_log, output_folder, spoc_map
            )
            if output_path:
                created.append(output_path)
                logging.info("Created state workbook: %s", output_path)

        if not created:
            raise ValueError("No state workbook was created.")

        print("Created state workbooks:")
        for output_path in created:
            print(f" - {output_path}")
        return 0

    except (ValueError, FileNotFoundError) as exc:
        logging.error("Configuration/runtime error: %s", exc)
        return 2
    except HttpError as exc:
        logging.error("Google API request failed: %s", exc)
        return 3
    except Exception as exc:
        logging.exception("Unexpected failure: %s", exc)
        return 1


if __name__ == "__main__":
    sys.exit(main())
